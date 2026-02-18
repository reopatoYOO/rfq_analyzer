"""
Template Mapper - Reads Excel template structure and maps spec fields.
"""

import logging
from pathlib import Path

import openpyxl

from models import TemplateField

logger = logging.getLogger(__name__)


class TemplateMapper:
    """Reads and analyzes the Excel template structure."""

    def __init__(self, template_path: str):
        self.template_path = template_path
        self.fields: list[TemplateField] = []

    def read_template(self) -> list[TemplateField]:
        """
        Read the Excel template and extract field definitions.

        The template is expected to have:
        - Column A: Specification type (spec name)
        - Column B: OEM requirement
        - Column C: LGE requirement

        Returns:
            List of TemplateField objects.
        """
        path = Path(self.template_path)
        if not path.exists():
            raise FileNotFoundError(f"Template file not found: {self.template_path}")

        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active  # Use the first sheet

        self.fields = []

        for row in range(1, ws.max_row + 1):
            spec_name = ws.cell(row=row, column=1).value  # Column A
            oem_value = ws.cell(row=row, column=2).value  # Column B
            lge_value = ws.cell(row=row, column=3).value  # Column C

            if spec_name is not None:
                field = TemplateField(
                    row=row,
                    col_spec_name=1,
                    col_oem_value=2,
                    col_lge_value=3,
                    spec_name=str(spec_name).strip(),
                    oem_value=str(oem_value).strip() if oem_value else "",
                    lge_value=str(lge_value).strip() if lge_value else "",
                )
                self.fields.append(field)

        wb.close()

        # Skip header row if it looks like a header
        if self.fields and self.fields[0].spec_name.lower() in (
            "specification type", "spec type", "item", "specification",
        ):
            self.header_row = self.fields[0]
            self.fields = self.fields[1:]  # Remove header from fields
        else:
            self.header_row = None

        logger.info(
            f"Read template: {len(self.fields)} spec fields from {path.name}"
        )
        return self.fields

    def get_spec_names(self) -> list[str]:
        """Get list of all specification names from the template."""
        return [f.spec_name for f in self.fields]
