"""
Excel Writer - Writes extracted specs into the Excel template with references.
"""

import logging
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import ExtractedSpec, MappingResult, SpecReference

logger = logging.getLogger(__name__)

# Confidence-based cell colors
CONFIDENCE_COLORS = {
    "high": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),     # Green
    "medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),   # Yellow
    "low": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),      # Red
}


def _confidence_level(confidence: float) -> str:
    """Map confidence score to level."""
    if confidence >= 0.8:
        return "high"
    elif confidence >= 0.5:
        return "medium"
    return "low"


class ExcelWriter:
    """Writes spec results into an Excel template with references and formatting."""

    def __init__(self, template_path: str, output_folder: str):
        self.template_path = template_path
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(parents=True, exist_ok=True)

    def write_results(
        self,
        mappings: list[MappingResult],
        unmatched_specs: list[ExtractedSpec],
        output_filename: str | None = None,
    ) -> str:
        """
        Write extracted specs into the template and create output Excel file.

        Creates three sheets:
        1. Spec Summary - Template with filled values + cell comments
        2. Reference - Source reference table
        3. Unmatched - Specs not mapped to template

        Args:
            mappings: List of spec-to-template mappings.
            unmatched_specs: Specs that couldn't be mapped.
            output_filename: Output file name (auto-generated if None).

        Returns:
            Path to the output file.
        """
        if not output_filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"RFQ_Spec_Result_{timestamp}.xlsx"

        output_path = self.output_folder / output_filename

        # Load template
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb.active
        ws.title = "Spec Summary"

        # Write mapped specs into template
        self._write_spec_summary(ws, mappings)

        # Create Reference sheet
        ws_ref = wb.create_sheet("Reference")
        self._write_reference_sheet(ws_ref, mappings)

        # Create Unmatched sheet
        ws_unmatched = wb.create_sheet("Unmatched")
        self._write_unmatched_sheet(ws_unmatched, unmatched_specs)

        # Save
        wb.save(str(output_path))
        wb.close()

        logger.info(f"Results saved to: {output_path}")
        return str(output_path)

    def _write_spec_summary(self, ws, mappings: list[MappingResult]):
        """Write spec values into the template sheet with comments and colors."""
        for mapping in mappings:
            field = mapping.template_field
            spec = mapping.extracted_spec
            ref = spec.reference

            # Write the OEM requirement value (Column B)
            cell = ws.cell(row=field.row, column=field.col_oem_value)
            cell.value = spec.value

            # Apply confidence-based color
            level = _confidence_level(mapping.match_confidence)
            cell.fill = CONFIDENCE_COLORS[level]

            # Add comment with source reference
            if ref:
                comment_text = (
                    f"Source: {ref.source_file}\n"
                    f"Location: {ref.page_label}\n"
                    f"Original: {ref.original_text[:200]}\n"
                    f"Confidence: {ref.confidence:.0%}"
                )
                cell.comment = Comment(comment_text, "RFQ Analyzer")
                cell.comment.width = 400
                cell.comment.height = 150

        logger.info(f"Wrote {len(mappings)} spec values to Spec Summary sheet")

    def _write_reference_sheet(self, ws, mappings: list[MappingResult]):
        """Create the Reference sheet with all source references."""
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = [
            "Spec Name",
            "Extracted Value",
            "Unit",
            "Condition",
            "Source File",
            "Location",
            "Original Text",
            "Translated Text",
            "Confidence",
        ]

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Write data
        for i, mapping in enumerate(mappings, start=2):
            spec = mapping.extracted_spec
            ref = spec.reference

            row_data = [
                spec.spec_name,
                spec.value,
                spec.unit,
                spec.condition,
                ref.source_file if ref else "",
                ref.page_label if ref else "",
                ref.original_text if ref else "",
                ref.translated_text if ref else "",
                f"{mapping.match_confidence:.0%}",
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Auto-width columns (approximate)
        col_widths = [25, 20, 10, 20, 30, 15, 50, 50, 12]
        for col, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        logger.info(f"Wrote {len(mappings)} references to Reference sheet")

    def _write_unmatched_sheet(self, ws, unmatched_specs: list[ExtractedSpec]):
        """Create the Unmatched sheet for specs not mapped to template."""
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = [
            "Spec Name",
            "Value",
            "Unit",
            "Condition",
            "Source File",
            "Location",
            "Original Text",
            "Confidence",
        ]

        # Write headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Write data
        for i, spec in enumerate(unmatched_specs, start=2):
            ref = spec.reference

            row_data = [
                spec.spec_name,
                spec.value,
                spec.unit,
                spec.condition,
                ref.source_file if ref else "",
                ref.page_label if ref else "",
                ref.original_text if ref else "",
                f"{spec.confidence:.0%}",
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Auto-width columns
        col_widths = [25, 20, 10, 20, 30, 15, 50, 12]
        for col, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        if unmatched_specs:
            logger.info(
                f"Wrote {len(unmatched_specs)} unmatched specs to Unmatched sheet"
            )
        else:
            # Write a "no unmatched specs" message
            ws.cell(row=2, column=1, value="No unmatched specifications found.")
