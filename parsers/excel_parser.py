"""
Excel Parser - Extracts text and tables from Excel files using openpyxl.
"""

import logging
from pathlib import Path

import openpyxl

from models import ParsedDocument, ParsedPage

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parses Excel files and extracts data per sheet."""

    def parse(self, file_path: str) -> ParsedDocument:
        """
        Parse an Excel file and return a ParsedDocument.

        Args:
            file_path: Path to the Excel file.

        Returns:
            ParsedDocument with pages (one per sheet) containing text and tables.
        """
        path = Path(file_path)
        doc = ParsedDocument(
            file_path=str(path.resolve()),
            file_name=path.name,
            file_type="xlsx",
        )

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)

            for i, sheet_name in enumerate(wb.sheetnames, start=1):
                ws = wb[sheet_name]

                # Read all rows as table data
                table_data = []
                text_parts = []

                for row in ws.iter_rows(
                    min_row=1,
                    max_row=ws.max_row,
                    max_col=ws.max_column,
                    values_only=True,
                ):
                    row_values = [
                        str(cell) if cell is not None else "" for cell in row
                    ]
                    # Skip completely empty rows
                    if any(v.strip() for v in row_values):
                        table_data.append(row_values)
                        text_parts.append(" | ".join(v for v in row_values if v.strip()))

                full_text = "\n".join(text_parts)

                parsed_page = ParsedPage(
                    page_number=i,
                    page_label=f'Sheet "{sheet_name}"',
                    text=full_text,
                    tables=[table_data] if table_data else [],
                )
                doc.pages.append(parsed_page)

            wb.close()
            logger.info(f"Parsed Excel: {path.name} ({len(doc.pages)} sheets)")

        except Exception as e:
            logger.error(f"Failed to parse Excel {path.name}: {e}")
            raise

        return doc
